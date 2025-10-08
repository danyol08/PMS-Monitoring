import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
import io

def generate_excel_report(data, contract_type=None):
    """Generate Excel report from contract data"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if contract_type == "hardware" or contract_type is None:
            if "hardware" in data and data["hardware"]:
                hw_df = pd.DataFrame(data["hardware"])
                hw_df.to_excel(writer, sheet_name="Hardware Contracts", index=False)
        
        if contract_type == "label" or contract_type is None:
            if "label" in data and data["label"]:
                label_df = pd.DataFrame(data["label"])
                label_df.to_excel(writer, sheet_name="Label Contracts", index=False)
        
        # Summary sheet
        summary_data = []
        if "hardware" in data:
            summary_data.append(["Hardware Contracts", len(data["hardware"])])
        if "label" in data:
            summary_data.append(["Label Contracts", len(data["label"])])
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data, columns=["Contract Type", "Count"])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
    
    output.seek(0)
    return output.getvalue()

def generate_pdf_report(data, contract_type=None):
    """Generate PDF report from contract data"""
    pass

def generate_service_history_excel(service_history_data):
    """Generate Excel report for service history with the new table format"""
    import pandas as pd
    import io
    
    # Create DataFrame with the new table format
    df_data = []
    for i, service in enumerate(service_history_data, 1):
        df_data.append({
            'NO': i,
            'COMPANY': service.get('company', ''),
            'LOCATION': service.get('location', ''),
            'MODEL': service.get('model', ''),
            'SERIAL': service.get('serial', ''),
            'DATE OF PMS': service.get('service_date', ''),
            'TECHNICAL MEMBER': service.get('technician', ''),
            'SALES': service.get('sales', ''),
            'SR': service.get('sr_number', ''),
            'SERVICE REPORT': service.get('service_report', '')
        })
    
    df = pd.DataFrame(df_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Service History', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Service History']
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()

def generate_service_history_pdf(service_history_data):
    """Generate PDF report for service history with the new table format"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("Service History Report", title_style)
    
    # Table data - Note: PDF table will be split into two parts due to width constraints
    # Part 1: Basic info
    table_data_1 = [['NO', 'COMPANY', 'LOCATION', 'MODEL', 'SERIAL', 'DATE OF PMS', 'TECHNICAL MEMBER']]
    
    for i, service in enumerate(service_history_data, 1):
        table_data_1.append([
            str(i),
            service.get('company', ''),
            service.get('location', ''),
            service.get('model', ''),
            service.get('serial', ''),
            service.get('service_date', ''),
            service.get('technician', '')
        ])
    
    # Part 2: Additional info
    table_data_2 = [['NO', 'SALES', 'SR', 'SERVICE REPORT']]
    
    for i, service in enumerate(service_history_data, 1):
        # Truncate service report for PDF display
        service_report = service.get('service_report', '')
        if len(service_report) > 100:
            service_report = service_report[:100] + '...'
        
        table_data_2.append([
            str(i),
            service.get('sales', ''),
            service.get('sr_number', ''),
            service_report
        ])
    
    # Create tables with styling
    table_style = TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    table_1.setStyle(table_style)
    table_2.setStyle(table_style)
    
    # Build PDF with both tables
    elements = [
        title, 
        Spacer(1, 20), 
        Paragraph("Basic Service Information", styles['Heading2']),
        Spacer(1, 10),
        table_1,
        Spacer(1, 30),
        Paragraph("Additional Service Details", styles['Heading2']),
        Spacer(1, 10),
        table_2
    ]
    doc.build(elements)
    
    buffer.seek(0)
    return buffer.getvalue()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph("Preventive Maintenance System Report", title_style))
    story.append(Spacer(1, 12))
    
    # Report date
    date_style = ParagraphStyle(
        'CustomDate',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1
    )
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
    story.append(Spacer(1, 20))
    
    # Hardware contracts table
    if contract_type == "hardware" or contract_type is None:
        if "hardware" in data and data["hardware"]:
            story.append(Paragraph("Hardware Contracts", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            hw_data = data["hardware"]
            if hw_data:
                # Prepare table data
                table_data = [["SQ", "End User", "Model", "Serial", "Status", "Branch", "Next PMS"]]
                
                for contract in hw_data[:50]:  # Limit to 50 records for PDF
                    table_data.append([
                        contract.get("sq", ""),
                        contract.get("end_user", ""),
                        contract.get("model", ""),
                        contract.get("serial", ""),
                        contract.get("status", ""),
                        contract.get("branch", ""),
                        contract.get("next_pms_schedule", "")[:10] if contract.get("next_pms_schedule") else ""
                    ])
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
    
    # Label contracts table
    if contract_type == "label" or contract_type is None:
        if "label" in data and data["label"]:
            story.append(Paragraph("Label Contracts", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            label_data = data["label"]
            if label_data:
                # Prepare table data
                table_data = [["SQ", "End User", "Part Number", "Serial", "Status", "Branch", "Next PMS"]]
                
                for contract in label_data[:50]:  # Limit to 50 records for PDF
                    table_data.append([
                        contract.get("sq", ""),
                        contract.get("end_user", ""),
                        contract.get("part_number", ""),
                        contract.get("serial", ""),
                        contract.get("status", ""),
                        contract.get("branch", ""),
                        contract.get("next_pms_schedule", "")[:10] if contract.get("next_pms_schedule") else ""
                    ])
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
    
    # Summary
    story.append(Paragraph("Summary", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    summary_data = []
    if "hardware" in data:
        summary_data.append(["Hardware Contracts", str(len(data["hardware"]))])
    if "label" in data:
        summary_data.append(["Label Contracts", str(len(data["label"]))])
    
    if summary_data:
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def calculate_next_maintenance(current_date, frequency):
    """Calculate next maintenance date based on frequency"""
    from datetime import timedelta
    
    if frequency == "monthly":
        return current_date + timedelta(days=30)
    elif frequency == "quarterly":
        return current_date + timedelta(days=90)
    elif frequency == "semi-annual":
        return current_date + timedelta(days=180)
    elif frequency == "yearly":
        return current_date + timedelta(days=365)
    else:
        return current_date + timedelta(days=30)  # Default to monthly

def generate_repairs_history_excel(repairs_data):
    """Generate Excel report for repairs history"""
    import pandas as pd
    import io
    
    # Create DataFrame
    df_data = []
    for i, repair in enumerate(repairs_data, 1):
        df_data.append({
            'NO': i,
            'SQ': repair.get('sq', ''),
            'DATE RECEIVED': repair.get('date_received', ''),
            'DATE COMPLETED': repair.get('repair_closed', ''),
            'COMPANY': repair.get('company_name', ''),
            'MODEL': repair.get('device_model', ''),
            'SERIAL': repair.get('serial_number', ''),
            'PART NUMBER': repair.get('part_number', ''),
            'RMA CASE': repair.get('rma_case', ''),
            'TECHNICIAN': repair.get('technician_notes', ''),
            'ACTION TAKEN': repair.get('action_taken', ''),
            'COMPLETION NOTES': repair.get('completion_notes', '')
        })
    
    df = pd.DataFrame(df_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Repairs History', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Repairs History']
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()




